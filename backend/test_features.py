import unittest
import os
import json
import sqlite3
import ipaddress
import openpyxl
from fastapi.testclient import TestClient

# Mock the Excel file path and DB path BEFORE importing application modules
# to avoid modifying the real files during tests.
os.environ["TESTING"] = "1"
TEST_EXCEL_FILE = 'test_ip_template.xlsx'
TEST_DB_FILE = 'test_ipnex.db'

# We have to patch the modules after import or right before
import database
database.DB_FILE = TEST_DB_FILE

import excel_sync
excel_sync.EXCEL_FILE = TEST_EXCEL_FILE

from main import app
from excel_sync import sync_excel_to_db, sync_db_to_excel, init_excel_if_not_exists
from database import init_db, get_db_connection

client = TestClient(app)

class TestIPNexFeatures(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        # Ensure fresh state
        if os.path.exists(TEST_DB_FILE):
            os.remove(TEST_DB_FILE)
        if os.path.exists(TEST_EXCEL_FILE):
            os.remove(TEST_EXCEL_FILE)
            
        init_db()

    @classmethod
    def tearDownClass(cls):
        if os.path.exists(TEST_DB_FILE):
            os.remove(TEST_DB_FILE)
        if os.path.exists(TEST_EXCEL_FILE):
            os.remove(TEST_EXCEL_FILE)

    def test_1_excel_sync(self):
        # 1. Init dummy excel file
        init_excel_if_not_exists()
        wb = openpyxl.load_workbook(TEST_EXCEL_FILE)
        ws = wb.active
        # Add a dummy device
        ws.append(["Test-PC", "192.168.1.50", "aa:bb:cc:dd:ee:ff", "online", "2026-05-01 10:00:00"])
        wb.save(TEST_EXCEL_FILE)
        
        # 2. Sync to DB
        sync_excel_to_db()
        
        # 3. Check if DB has the device
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM devices WHERE ip_address = '192.168.1.50'")
        device = cursor.fetchone()
        conn.close()
        
        self.assertIsNotNone(device, "Device from Excel should be synced to DB")
        self.assertEqual(device['hostname'], "Test-PC")
        self.assertEqual(device['mac_address'], "aa:bb:cc:dd:ee:ff")

    def test_2_unused_ips_endpoint(self):
        # 1. We know 192.168.1.50 is used now. Check the endpoint.
        response = client.get("/api/network/unused?subnet=192.168.1.0/24")
        self.assertEqual(response.status_code, 200)
        
        data = response.json()
        self.assertEqual(data["subnet"], "192.168.1.0/24")
        self.assertIn("unused_ips", data)
        
        # Total IPs in /24 is 256. 192.168.1.50 is used, so it shouldn't be in unused_ips
        self.assertNotIn("192.168.1.50", data["unused_ips"])
        self.assertIn("192.168.1.51", data["unused_ips"])
        # Network and broadcast are included because strict=False and we just do network.hosts() which might return 254 hosts (excluding network/broadcast).
        self.assertTrue(data["total_unused"] > 0)

    def test_3_sync_db_to_excel(self):
        # 1. Add a device via DB manually
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO devices (hostname, ip_address, mac_address, status, is_reserved, first_seen, last_seen)
            VALUES ('DB-PC', '192.168.1.55', '11:22:33:44:55:66', 'online', TRUE, '2026-05-01 10:00:00', '2026-05-01 10:00:00')
            ON CONFLICT (mac_address) DO NOTHING
        ''')
        conn.commit()
        conn.close()
        
        # 2. Sync to Excel
        sync_db_to_excel()
        
        # 3. Read Excel and Verify
        wb = openpyxl.load_workbook(TEST_EXCEL_FILE)
        ws = wb.active
        
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Should now have 2 rows
        self.assertEqual(len(rows), 2)
        ip_addresses = [str(r[1]) for r in rows]
        self.assertIn("192.168.1.50", ip_addresses)
        self.assertIn("192.168.1.55", ip_addresses)

if __name__ == '__main__':
    unittest.main()
