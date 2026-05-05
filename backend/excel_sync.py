import os
import openpyxl
from datetime import datetime
from database import get_db_connection
from logger import get_logger

logger = get_logger(__name__)

TEMPLATE_FILE = r'd:\SAM\netX\ip_template.xlsx'

def get_current_excel_file():
    month_yr = datetime.now().strftime("%b_%y")
    return fr'd:\SAM\netX\IP_{month_yr}.xlsx'

def get_active_read_file():
    target = get_current_excel_file()
    return target if os.path.exists(target) else TEMPLATE_FILE

def init_excel_if_not_exists():
    pass # Disabled as we rely on the template file being present

def sync_excel_to_db():
    read_file = get_active_read_file()
    if not os.path.exists(read_file):
        logger.error(f"Error: {read_file} not found!")
        return
        
    try:
        wb = openpyxl.load_workbook(read_file)
        ws = wb.active
        
        header_row_idx = 1
        headers = []
        for row_idx in range(1, 10):
            row_vals = [str(cell.value).strip() if cell.value else "" for cell in ws[row_idx]]
            if any("IP" in v.upper() for v in row_vals):
                headers = row_vals
                header_row_idx = row_idx
                break
                
        if not headers:
            return
            
        ip_name = next((h for h in headers if "IP" in h.upper()), None)
        control_name = next((h for h in headers if "CONTROL" in h.upper()), None)
        process_name = next((h for h in headers if "PROCESS" in h.upper()), None)
        model_name = next((h for h in headers if "MODEL" in h.upper()), None)
        mac_name = next((h for h in headers if "MAC" in h.upper()), None)
        
        if not ip_name:
            return
            
        ip_idx = headers.index(ip_name)
        control_idx = headers.index(control_name) if control_name else -1
        process_idx = headers.index(process_name) if process_name else -1
        model_idx = headers.index(model_name) if model_name else -1
        mac_idx = headers.index(mac_name) if mac_name else -1
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        count = 0
        current_model = ""
        current_process = ""
        
        for row in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
            if len(row) <= ip_idx or not row[ip_idx]:
                continue
            
            ip = str(row[ip_idx]).strip()
            hostname = str(row[control_idx]).strip() if control_idx >= 0 and row[control_idx] else ""
            
            raw_process = str(row[process_idx]).strip() if process_idx >= 0 and row[process_idx] else ""
            if raw_process: current_process = raw_process
            process_val = current_process
            
            raw_model = str(row[model_idx]).strip() if model_idx >= 0 and row[model_idx] else ""
            if raw_model: current_model = raw_model
            model_val = current_model
            
            mac = str(row[mac_idx]).strip().lower() if mac_idx >= 0 and row[mac_idx] else f"unknown-{ip}"
            
            cursor.execute("SELECT * FROM devices WHERE ip_address = %s", (ip,))
            if not cursor.fetchone():
                cursor.execute('''
                    INSERT INTO devices (hostname, model, process, ip_address, mac_address, status, is_reserved, first_seen, last_seen)
                    VALUES (%s, %s, %s, %s, %s, 'online', TRUE, %s, %s)
                ''', (hostname, model_val, process_val, ip, mac, now_str, now_str))
                count += 1
            else:
                cursor.execute('''
                    UPDATE devices SET hostname = %s, model = %s, process = %s WHERE ip_address = %s
                ''', (hostname, model_val, process_val, ip))
        conn.commit()
        conn.close()
    except Exception as e:
        logger.error(f"Error syncing from excel: {e}", exc_info=True)

def sync_db_to_excel():
    read_file = get_active_read_file()
    if not os.path.exists(read_file):
        return
        
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT hostname, model, process, ip_address, mac_address, status, last_seen FROM devices")
        db_devices = cursor.fetchall()
        conn.close()

        ip_map = {d['ip_address']: d for d in db_devices}

        wb = openpyxl.load_workbook(read_file)
        ws = wb.active
        
        header_row_idx = 1
        headers = []
        for row_idx in range(1, 10):
            row_vals = [str(cell.value).strip() if cell.value else "" for cell in ws[row_idx]]
            if any("IP" in v.upper() for v in row_vals):
                headers = row_vals
                header_row_idx = row_idx
                break
                
        if not headers:
            return

        ip_name = next((h for h in headers if "IP" in h.upper()), None)
        control_name = next((h for h in headers if "CONTROL" in h.upper()), None)
        process_name = next((h for h in headers if "PROCESS" in h.upper()), None)
        model_name = next((h for h in headers if "MODEL" in h.upper()), None)
        mac_name = next((h for h in headers if "MAC" in h.upper()), None)
        status_name = next((h for h in headers if "STATUS" in h.upper()), None)
        seen_name = next((h for h in headers if "SEEN" in h.upper() or "DATE" in h.upper()), None)
        
        ip_idx = headers.index(ip_name) + 1
        max_col = len(headers)
        
        if not control_name:
            max_col += 1; ws.cell(row=header_row_idx, column=max_col, value="Control"); control_idx = max_col
        else: control_idx = headers.index(control_name) + 1

        if not process_name:
            max_col += 1; ws.cell(row=header_row_idx, column=max_col, value="Process"); process_idx = max_col
        else: process_idx = headers.index(process_name) + 1

        if not model_name:
            max_col += 1; ws.cell(row=header_row_idx, column=max_col, value="Model"); model_idx = max_col
        else: model_idx = headers.index(model_name) + 1

        if not mac_name:
            max_col += 1
            ws.cell(row=header_row_idx, column=max_col, value="MAC Address")
            mac_idx = max_col
        else:
            mac_idx = headers.index(mac_name) + 1

        if not status_name:
            max_col += 1
            ws.cell(row=header_row_idx, column=max_col, value="Status")
            status_idx = max_col
        else:
            status_idx = headers.index(status_name) + 1

        if not seen_name:
            max_col += 1
            ws.cell(row=header_row_idx, column=max_col, value="Last Seen")
            seen_idx = max_col
        else:
            seen_idx = headers.index(seen_name) + 1

        existing_ips = set()
        max_row = ws.max_row
        def safe_write(r, c, val):
            if c:
                cell = ws.cell(row=r, column=c)
                # Avoid writing to secondary cells in a merged range
                if type(cell).__name__ != 'MergedCell':
                    cell.value = str(val) if val is not None else ""

        for row in range(header_row_idx+1, max_row + 1):
            ip_val = ws.cell(row=row, column=ip_idx).value
            if not ip_val: continue
            ip_str = str(ip_val).strip()
            existing_ips.add(ip_str)
            
            if ip_str in ip_map:
                d = ip_map[ip_str]
                safe_write(row, control_idx, d['hostname'])
                safe_write(row, process_idx, d['process'])
                safe_write(row, model_idx, d['model'])
                safe_write(row, mac_idx, d['mac_address'])
                safe_write(row, status_idx, d['status'])
                safe_write(row, seen_idx, d['last_seen'])
                
        # Append new local IPs not registered in Excel
        seen_ip = existing_ips
        for d in db_devices:
            if d['ip_address'] not in seen_ip:
                new_row = ["" for _ in range(max_col)]
                if control_idx: new_row[control_idx-1] = d['hostname']
                if process_idx: new_row[process_idx-1] = d['process']
                if model_idx: new_row[model_idx-1] = d['model']
                if ip_idx: new_row[ip_idx-1] = d['ip_address']
                if mac_idx: new_row[mac_idx-1] = d['mac_address']
                if status_idx: new_row[status_idx-1] = d['status']
                if seen_idx: new_row[seen_idx-1] = str(d['last_seen'])
                ws.append(new_row)
            
        # Method 2: Write Unused IPs into 'Address (XXX)' columns
        import re
        address_cols = {}
        for idx, h in enumerate(headers):
            match = re.search(r'ADDRESS\s*\(\s*(\d+)\s*\)', h.upper())
            if match:
                address_cols[idx + 1] = match.group(1)
                
        if address_cols:
            all_used_ips = set(ip_map.keys())
            
            # Predict the first two octets of the network (e.g. 192.168)
            base_prefix = "192.168"
            for u_ip in all_used_ips:
                parts = u_ip.split('.')
                if len(parts) == 4:
                    base_prefix = f"{parts[0]}.{parts[1]}"
                    break
                    
            for col_idx, third_oct in address_cols.items():
                subnet_prefix = f"{base_prefix}.{third_oct}."
                
                # Retrieve current max_row to erase existing old static empty lists
                current_max = ws.max_row
                for r in range(header_row_idx + 1, current_max + 1):
                    safe_write(r, col_idx, "")
                
                # Calculate Unused IPs 1 to 254
                unused_ips = []
                for i in range(1, 255):
                    cand = f"{subnet_prefix}{i}"
                    if cand not in all_used_ips:
                        unused_ips.append(cand)
                        
                # Write dynamically to the column
                write_r = header_row_idx + 1
                for cand in unused_ips:
                    safe_write(write_r, col_idx, cand)
                    write_r += 1
            
        output_file = get_current_excel_file()
        wb.save(output_file)
        logger.info(f"Data saved to {output_file}")
    except Exception as e:
        logger.error(f"Error syncing to excel: {e}", exc_info=True)
